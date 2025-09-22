import re
from datetime import date, datetime

import openpyxl
from django.core.exceptions import ValidationError


def extract_players_from_excel(
    excel_file, header_row=24, start_data_row=25, default_position="MID"
):
    """
    Extrae datos de jugadores del archivo Excel

    Args:
        excel_file: Archivo Excel subido
        header_row: Fila donde están los encabezados (1-indexed)
        start_data_row: Fila donde empiezan los datos (1-indexed)
        default_position: Posición por defecto para los jugadores
    """
    try:
        # Cargar el workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        # Obtener la primera hoja
        worksheet = workbook.active

        # Leer encabezados para mapear columnas
        headers = {}
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value:
                header_name = str(cell_value).strip().upper()
                headers[col] = header_name

        print(f"Encabezados encontrados: {headers}")

        # Mapear columnas según los encabezados encontrados
        column_mapping = find_column_mapping(headers)

        # Extraer datos de jugadores
        players_data = []

        for row in range(start_data_row, worksheet.max_row + 1):
            # Verificar si la fila tiene datos
            if is_empty_row(worksheet, row, column_mapping):
                continue

            try:
                player_data = extract_player_from_row(
                    worksheet, row, column_mapping, default_position
                )
                if player_data:
                    players_data.append(player_data)

            except Exception as e:
                print(f"Error procesando fila {row}: {str(e)}")
                continue

        return players_data

    except Exception as e:
        raise ValidationError(f"Error al procesar el archivo Excel: {str(e)}")


def find_column_mapping(headers):
    """
    Mapea las columnas del Excel según los encabezados encontrados
    """
    mapping = {}

    for col, header in headers.items():
        header = header.upper().strip()

        # Mapear apellidos
        if any(word in header for word in ["APELLIDO", "SURNAME", "LAST"]):
            mapping["apellido"] = col

        # Mapear nombres
        elif any(word in header for word in ["NOMBRE", "NAME", "FIRST"]):
            mapping["nombre"] = col

        # Mapear DNI
        elif any(word in header for word in ["DNI", "DOCUMENTO", "ID", "CEDULA"]):
            mapping["dni"] = col

        # Mapear fecha de nacimiento
        elif any(
            word in header for word in ["FECHA", "NAC", "BIRTH", "BORN", "NACIMIENTO"]
        ):
            mapping["fecha_nacimiento"] = col

        # Mapear número de jugador
        elif header.isdigit() or any(
            word in header for word in ["NUM", "NUMERO", "NUMBER", "#"]
        ):
            if "numero" not in mapping:  # Solo tomar el primero
                mapping["numero"] = col

        # Mapear celular/teléfono
        elif any(word in header for word in ["CELULAR", "TELEFONO", "PHONE", "TEL"]):
            mapping["celular"] = col

        # Mapear promoción
        elif any(word in header for word in ["PROMO", "PROMOCION", "CLASS", "YEAR"]):
            mapping["promocion"] = col
        
        # Mapear oficio/profesión
        elif any(word in header for word in ["OFICIO", "PROFESION", "OCCUPATION", "JOB"]):
            mapping["oficio"] = col

    print(f"Mapeo de columnas: {mapping}")
    return mapping


def is_empty_row(worksheet, row, column_mapping):
    """
    Verifica si una fila está vacía o contiene solo datos irrelevantes
    """
    important_cols = ["apellido", "nombre", "dni"]

    for col_name in important_cols:
        if col_name in column_mapping:
            cell_value = worksheet.cell(row=row, column=column_mapping[col_name]).value
            if cell_value and str(cell_value).strip():
                return False

    return True


def extract_player_from_row(worksheet, row, column_mapping, default_position):
    """
    Extrae los datos de un jugador desde una fila específica
    """
    player_data = {}

    # Apellido (obligatorio)
    if "apellido" in column_mapping:
        apellido = worksheet.cell(row=row, column=column_mapping["apellido"]).value
        if not apellido:
            return None
        player_data["last_name"] = str(apellido).strip()
    else:
        return None

    # Nombre (obligatorio)
    if "nombre" in column_mapping:
        nombre = worksheet.cell(row=row, column=column_mapping["nombre"]).value
        if not nombre:
            return None
        player_data["first_name"] = str(nombre).strip()
    else:
        return None

    # DNI
    if "dni" in column_mapping:
        dni = worksheet.cell(row=row, column=column_mapping["dni"]).value
        if dni:
            # Limpiar DNI (quitar puntos, espacios, etc.)
            dni_str = str(dni).replace(".", "").replace(" ", "").replace(",", "")
            # Extraer solo números
            dni_clean = re.sub(r"[^\d]", "", dni_str)
            if len(dni_clean) >= 7:
                player_data["dni"] = dni_clean[:8]  # Máximo 8 dígitos
            else:
                player_data["dni"] = None
        else:
            player_data["dni"] = None

    # Fecha de nacimiento
    if "fecha_nacimiento" in column_mapping:
        fecha_nac = worksheet.cell(
            row=row, column=column_mapping["fecha_nacimiento"]
        ).value
        if fecha_nac:
            player_data["birth_date"] = parse_date_from_excel(fecha_nac)
        else:
            player_data["birth_date"] = date(2000, 1, 1)  # Fecha por defecto
    else:
        player_data["birth_date"] = date(2000, 1, 1)

    # Número de camiseta
    if "numero" in column_mapping:
        numero = worksheet.cell(row=row, column=column_mapping["numero"]).value
        if numero:
            try:
                player_data["jersey_number"] = str(int(numero))
            except:
                player_data["jersey_number"] = None
        else:
            player_data["jersey_number"] = None
    else:
        player_data["jersey_number"] = None

    # Posición (usar la por defecto)
    player_data["position"] = default_position

    # Información adicional para debugging
    player_data["row_number"] = row

    # Celular/Teléfono
    if "celular" in column_mapping:
        celular = worksheet.cell(row=row, column=column_mapping["celular"]).value
        if celular:
            player_data["phone"] = str(celular).strip()
        else:
            player_data["phone"] = ""
    else:
        player_data["phone"] = ""
        
    # Promoción
    if "promocion" in column_mapping:
        promocion = worksheet.cell(row=row, column=column_mapping["promocion"]).value
        if promocion:
            try:
                player_data["promo"] = int(promocion)
            except:
                player_data["promo"] = None
        else:
            player_data["promo"] = None
    else:
        player_data["promo"] = None
        
    # Oficio/Profesión
    if "oficio" in column_mapping:
        oficio = worksheet.cell(row=row, column=column_mapping["oficio"]).value
        if oficio:
            player_data["profession"] = str(oficio).strip()
        else:
            player_data["profession"] = ""
    else:
        player_data["profession"] = ""  
        
    return player_data


def parse_date_from_excel(date_value):
    """
    Convierte un valor de fecha de Excel a objeto date de Python
    """
    try:
        if isinstance(date_value, datetime):
            return date_value.date()
        elif isinstance(date_value, date):
            return date_value
        elif isinstance(date_value, str):
            # Intentar parsear string de fecha
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
                try:
                    return datetime.strptime(date_value, fmt).date()
                except ValueError:
                    continue

        # Si todo falla, fecha por defecto
        return date(2000, 1, 1)

    except:
        return date(2000, 1, 1)


def auto_assign_jersey_numbers(players_data, team):
    """
    Asigna números de camiseta automáticamente para jugadores que no los tienen
    """
    existing_numbers = set(team.players.values_list("jersey_number", flat=True))
    next_number = 1

    for player_data in players_data:
        if not player_data["jersey_number"]:
            while str(next_number) in existing_numbers:
                next_number += 1
            player_data["jersey_number"] = str(next_number)
            existing_numbers.add(str(next_number))
            next_number += 1


def auto_assign_dni(players_data, team):
    """
    Asigna DNI automáticamente para jugadores que no los tienen
    """
    existing_dnis = set(team.players.values_list("dni", flat=True))
    base_dni = 40000000

    for player_data in players_data:
        if not player_data["dni"]:
            while str(base_dni) in existing_dnis:
                base_dni += 1
            player_data["dni"] = str(base_dni)
            existing_dnis.add(str(base_dni))
            base_dni += 1
